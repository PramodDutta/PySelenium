# username, password
#  admin, admin
#  admin admin123
# admin password123
# contact+augg@thetestingacdey.com, Wingify@123
import time

import pytest
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.by import By


# Read the Excel the get the data

def get_test_data():
    workbook = load_workbook("testdata.xlsx")
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(row)
    return data


@pytest.fixture
def driver():
    driver = webdriver.Chrome()
    driver.get("http://app.vwo.com")
    driver.maximize_window()
    yield driver
    driver.quit()
    # Read a File, Read from DB, Create Webdriver


@pytest.mark.parametrize("username, password, result", get_test_data())
def test_vwo_login(driver, username, password, result):
    email = driver.find_element(By.ID, "login-username")
    email.send_keys(username)

    pass_word = driver.find_element(By.ID, "login-password")
    pass_word.send_keys(password)

    driver.find_element(By.ID, "js-login-btn").click()
    time.sleep(3)
    print(username, password, driver.current_url)
    if result == "fail":
        error_msg= driver.find_element(By.ID,"js-notification-box-msg").text
        assert error_msg in "Your email, password, IP address or location did not match"
    else:
        assert "'https://app.vwo.com/#/dashboard'" in driver.current_url
