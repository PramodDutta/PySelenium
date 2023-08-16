import pytest
from selenium import webdriver
from openpyxl import load_workbook


# pip install pytest openpyxl pytest-excel
def get_test_data():
    workbook = load_workbook("testdata.xlsx")
    sheet = workbook.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):  # Start from the second row to skip headers
        data.append(row)
    return data

@pytest.fixture
def setup_teardown():
    driver = webdriver.Chrome()
    driver.get("https://app.vwo.com")  # Replace with your website URL
    driver.maximize_window()
    yield driver
    driver.quit()


@pytest.mark.parametrize("username, password", get_test_data())
def test_login(setup_teardown, username, password):
    driver = setup_teardown
    print(username, password)

    # Add your assertions or validation steps here
    # For example, assert that a successful login redirects to the dashboard page

    # Wait for a brief moment to see the action
    driver.implicitly_wait(5)
